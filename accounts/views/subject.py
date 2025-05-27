from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse

from accounts import models
from utils.bootstrap import BootStrapModelForm
from utils.pagination import Pagination
# Create your views here.


class SubjectModelForm(BootStrapModelForm):
    class Meta:
        model = models.Subject
        fields = "__all__"


def subject_list(request):
    '''学科列表'''
    form = SubjectModelForm()
    # if request.method == 'GET':
    #     queryset = models.Subject.objects.all()
    queryset = models.Subject.objects.all()
    page_object = Pagination(request, queryset)

    contxt = {
        'form': form,
        'queryset': page_object.page_queryset,
        'page_string': page_object.html(),
    }
    return render(request, 'subject_list.html', contxt)


@csrf_exempt
def subject_add(request):
    '''学科添加(Ajax请求)'''
    form = SubjectModelForm(request.POST)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({'status': False, 'error': form.errors})


def subject_delete(request):
    '''学科删除'''
    uid = request.GET.get('uid')
    exists = models.Subject.objects.filter(id=uid).exists()
    if not exists:
        return JsonResponse({'status': False, 'error': '删除失败，数据不存在'})
    models.Subject.objects.filter(id=uid).delete()
    return JsonResponse({'status': True})


def subject_detail(request):
    '''根据ID获取学科详情'''
    uid = request.GET.get('uid')
    row_dict = models.Subject.objects.filter(id=uid).values('title').first()
    if not row_dict:
        return JsonResponse({"status": False, 'error': "数据不存在"})
    result = {
        'status': True,
        'data': row_dict,
    }
    return JsonResponse(result)


@csrf_exempt
def subject_edit(request):
    '''编辑学科'''
    uid = request.GET.get('uid')
    row_object = models.Subject.objects.filter(id=uid).first()
    if not row_object:
        return JsonResponse({"status": False, 'error': "数据不存在,请刷新重试"})

    form = SubjectModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({'status': False, 'error': form.errors})


def subject_multi(request):
    '''批量上传 (excel)'''
    from openpyxl import load_workbook
    # django.core.files.uploadedfile.InMemoryUploadedFile

    # 1.获取用户上传的对象文件
    file_object = request.FILES.get("exc")
    # print(type(file_object))

    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value

        exists = models.Subject.objects.filter(title=text).exists()
        if not exists:
            models.Subject.objects.create(title=text)

    # with open(file_object.title, mode='wb') as f:
    #     for chunk in file_object:
    #         f.write(chunk)

    return redirect("/accounts/subject/list/")
