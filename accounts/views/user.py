from django.shortcuts import render, redirect
from django import forms
from django.core.exceptions import ValidationError
from django.contrib import messages
from openpyxl import load_workbook
# from django.contrib.auth.hashers import make_password 用的直接封装好的set_password
from django.db.models import Q

from accounts import models
from utils.pagination import Pagination
from utils.bootstrap import BootStrapModelForm
from accounts.forms import ImportUsersForm
from utils.user_decorator import admin_or_superuser_required, superuser_required


@admin_or_superuser_required
def user_list(request):
    """用户列表"""
    # 获取可选数据
    subjects = models.Subject.objects.all()
    departments = models.Department.objects.all()
    
    # 初始化查询条件
    subject_id = request.GET.get('subject')
    department_id = request.GET.get('department')
    name = request.GET.get('name')
    
    # 构建查询条件
    query = Q()
    if subject_id and subject_id != 'all' :
        query &= Q(subject_id=subject_id)
    if department_id and department_id != 'all':
        query &= Q(department_id=department_id)
    if name:
        query &= Q(name__contains=name)
    
    # 应用查询条件
    queryset = models.UserInfo.objects.filter(query)
    # 分页
    page_object = Pagination(request, queryset)
    
    context = {
        'queryset': page_object.page_queryset,
        'page_string': page_object.html(),
        'subjects': subjects,
        'departments': departments,
        'seclect_subject': subject_id if subject_id else 'all', # 默认为all
        'seclect_department': department_id if department_id else 'all',
        'seclect_name': name if name else '',
    }

    return render(request, 'user_list.html', context)


class UserModelForm(BootStrapModelForm):
    confirm_password = forms.CharField(
        label='确认密码',
        widget=forms.PasswordInput(render_value=True),
    )

    class Meta:
        model = models.UserInfo
        fields = ['username', 'name', 'gender', 'phone', 'department', 'subject', 'password']
        widgets = {
            'password': forms.PasswordInput(render_value=True),
        }

    def clean_password(self):
        pwd = self.cleaned_data.get('password')
        return pwd

    def clean_confirm_password(self):
        pwd = self.cleaned_data.get('password')
        confirm = self.cleaned_data.get('confirm_password')
        if confirm != pwd and pwd is not None:
            raise ValidationError("两次输入的密码不一致")
        return confirm

    def save(self, commit=True):
        instance = super().save(commit=False)
        pwd = self.cleaned_data.get('password')
        if pwd:
            instance.set_password(pwd)
        if commit:
            instance.save()
        return instance


@superuser_required
def user_add(request):
    """添加用户"""
    title = '新建用户'
    if request.method == 'GET':
        form = UserModelForm()
        return render(request, 'change.html', {'form': form, 'title': title})

    form = UserModelForm(data=request.POST)
    if form.is_valid():
        form.save()
        return redirect('accounts:user_list')

    return render(request, 'change.html', {'form': form, 'title': title})


class UserEditModelForm(BootStrapModelForm):
    class Meta:
        model = models.UserInfo
        fields = ['username', 'name', 'gender', 'phone', 'department', 'subject']


@superuser_required
def user_edit(request, nid):
    """编辑用户"""
    title = '编辑用户'
    row_object = models.UserInfo.objects.filter(id=nid).first()
    if not row_object:
        return render(request, 'error.html', {"msg": "数据不存在"})

    if request.method == 'GET':
        form = UserEditModelForm(instance=row_object)
        return render(request, 'change.html', {'form': form, 'title': title})
    form = UserEditModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('accounts:user_list')
    return render(request, 'change.html', {'form': form, 'title': title})


@superuser_required
def user_delete(request):
    """删除用户"""
    nid = request.GET.get('nid')
    models.UserInfo.objects.filter(id=nid).delete()
    return redirect('accounts:user_list')


class UserRestModelForm(BootStrapModelForm):
    confirm_password = forms.CharField(
        label='确认密码',
        widget=forms.PasswordInput(render_value=True),
    )

    class Meta:
        model = models.UserInfo
        fields = ['password', 'confirm_password']
        widgets = {
            'password': forms.PasswordInput(render_value=True),
        }

    def clean_password(self):
        pwd = self.cleaned_data.get('password')
        # 检查新密码是否与旧密码相同
        if self.instance.pk and self.instance.check_password(pwd):
            raise ValidationError('密码不能与之前的相同')
        return pwd

    def clean_confirm_password(self):
        pwd = self.cleaned_data.get('password')
        confirm = self.cleaned_data.get('confirm_password')
        if confirm != pwd and pwd is not None:
            raise ValidationError("两次输入的密码不一致")
        return confirm

    def save(self, commit=True):
        instance = super().save(commit=False)
        pwd = self.cleaned_data.get('password')
        if pwd:
            instance.set_password(pwd)
        if commit:
            instance.save()
        return instance


@superuser_required
def user_reset(request, nid):
    """重置密码"""
    row_object = models.UserInfo.objects.filter(id=nid).first()
    if not row_object:
        return render(request, 'error.html', {"msg": "数据不存在"})

    title = '重置密码 - {}'.format(row_object.username)
    if request.method == 'GET':
        form = UserRestModelForm()
        return render(request, 'change.html', {'form': form, 'title': title})

    form = UserRestModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return redirect('accounts:user_list')
    return render(request, 'change.html', {'form': form, 'title': title})


@superuser_required
def import_users(request):
    if request.method == 'POST':
        form = ImportUsersForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']

            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, "错误：只支持 .xlsx 或 .xls 格式文件")
                return render(request, 'import_users.html', {'form': form})

            try:
                wb = load_workbook(excel_file)
                ws = wb.active

                success_count = 0
                skip_count = 0
                skip_list = []
                error_list = []

                subjects = {subject.title: subject for subject in models.Subject.objects.all()}
                departments = {dept.title: dept for dept in models.Department.objects.all()}

                for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):
                        continue

                    try:
                        username, name, subject_name, dept_name, phone = row[:5]

                        if models.UserInfo.objects.filter(username=username).exists():
                            skip_count += 1
                            skip_list.append(f"{username}（已存在）")
                            continue

                        subject_instance = subjects.get(subject_name)
                        if not subject_instance:
                            subject_instance = None

                        dept_instance = departments.get(dept_name)
                        if not dept_instance:
                            dept_instance = None

                        password = '123456'

                        user = models.UserInfo(
                            username=username,
                            name=name,
                            subject=subject_instance,
                            department=dept_instance,
                            phone=phone
                        )
                        user.set_password(password)
                        user.save()

                        success_count += 1

                    except Exception as e:
                        error_list.append(f"第 {row_index} 行错误: {str(e)}")
                        continue

                result_msg = f"导入完成！成功: {success_count}条"
                if skip_count > 0:
                    result_msg += f", 跳过: {skip_count}条"
                if skip_list:
                    messages.warning(request, f"跳过的用户: {', '.join(skip_list[:5])}" + ("..." if len(skip_list) > 5 else ""))
                if error_list:
                    messages.error(request, f"错误行: {', '.join(error_list[:3])}" + ("..." if len(error_list) > 3 else ""))

                messages.success(request, result_msg)
                return render(request, 'import_result.html')

            except Exception as e:
                messages.error(request, f"处理文件时出错: {str(e)}")

    else:
        form = ImportUsersForm()

    return render(request, 'import_result.html', {'form': form})
