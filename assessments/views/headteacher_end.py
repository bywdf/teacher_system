from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from django.db.models import Q
from django.db import transaction

from utils.pagination import Pagination
from utils.bootstrap import BootStrapModelForm

from assessments.models import HeadTeacherFinalAssess, AssessDepart, Semester, TermType
from accounts.models import UserInfo, Subject
from utils.user_decorator import superuser_required


class AssessModelForm(BootStrapModelForm):
    class Meta:
        model = HeadTeacherFinalAssess
        # 字段，所有字段
        fields = '__all__'


def headteacher_end_list(request):
    # 获取所有可选数据
    semesters = Semester.objects.order_by('-id')
    term_types = TermType.objects.all()
    assess_departs = AssessDepart.objects.all()
    subjects = Subject.objects.all()

    # 初始化查询条件
    semester_id = request.GET.get('semester')
    term_type_id = request.GET.get('term_type')
    assess_depart_id = request.GET.get('assess_depart')
    teacher_name = request.GET.get('teacher_name')
    subject_id = request.GET.get('subject')  # 新增学科参数

    # 构建查询条件
    query = Q()
    if semester_id and semester_id != 'all':
        query &= Q(semester_id=semester_id)
    if term_type_id and term_type_id != 'all':
        query &= Q(term_type_id=term_type_id)
    if assess_depart_id and assess_depart_id != 'all':
        query &= Q(assess_depart_id=assess_depart_id)
    if teacher_name:
        query &= Q(teacher__name__icontains=teacher_name)
    if subject_id and subject_id != 'all':
        query &= Q(teacher__subject_id=subject_id)
        
    # 权限控制：普通用户只能查看自己的记录
    if not (request.user.is_superuser or request.user.groups.filter(name='管理员').exists()):
        query &= Q(teacher=request.user)

    # 应用查询条件
    queryset = HeadTeacherFinalAssess.objects.filter(
        query).order_by('id')

    page_object = Pagination(request, queryset)

    content = {
        'queryset': page_object.page_queryset,
        'page_string': page_object.html(),
        'semesters': semesters,
        'term_types': term_types,
        'assess_departs': assess_departs,
        'subjects': subjects,
        'selected_semester': semester_id if semester_id else 'all',
        'selected_term_type': term_type_id if term_type_id else 'all',
        'selected_assess_depart': assess_depart_id if assess_depart_id else 'all',
        'teacher_name': teacher_name if teacher_name else '',
        'selected_subject': subject_id if subject_id else 'all',

    }
    return render(request, 'headteacher_end_list.html', content)


@superuser_required
def headteacher_end_delete(request):
    """删除"""
    nid = request.GET.get('nid')
    HeadTeacherFinalAssess.objects.filter(id=nid).delete()
    return redirect('assessments:headteacher_end_list')


@superuser_required
def headteacher_end_edit(request, pk):
    # 获取要编辑的对象，若不存在则返回404
    instance = get_object_or_404(HeadTeacherFinalAssess, pk=pk)
    # 创建表单实例，绑定现有数据
    form = AssessModelForm(request.POST or None, instance=instance)

    if request.method == 'POST':
        if form.is_valid():
            # 保存更新（可在此处添加额外逻辑，如权限检查、计算字段等）
            form.save()
            return redirect('assessments:headteacher_end_list')  # 重定向到列表页
        # 若表单验证失败，保留错误信息并重新渲染页面

    # 渲染编辑页面，传递表单和对象
    context = {
        'form': form,
        'title': '考核记录',
        'instance': instance,
        'show_class_field':True,
        'show_manage_score':True,
        'show_safety_score':True,
        'show_class_score':True,
    }
    return render(request, 'assess_change.html', context)


@superuser_required
def headteacher_end_add(request):
    """添加"""
    form = AssessModelForm()
    # 获取所有教师数据
    teachers = UserInfo.objects.all()
    if request.method == 'POST':
        form = AssessModelForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('assessments:headteacher_end_list')
        
    content = {
        'form': form,
        'title': '添加考核记录',
        'show_class_field':True,
        'show_manage_score':True,
        'show_safety_score':True,
        'show_class_score':True,
    }
    
    return render(request, 'assess_change.html', content)


# 下面是批量导入需要的功能
@superuser_required
@transaction.atomic
def headteacher_end_import(request):
    """批量导入考核成绩，支持一个教师担任多个班级班主任"""
    if request.method == "POST":
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "请选择Excel文件")
            return redirect('assessments:headteacher_end_list')

        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            errors = []
            success_count = 0
            created_count = 0
            updated_count = 0

            # 学期映射字典
            semester_map = {}
            for sem in Semester.objects.all():
                key = f"{sem.year}{sem.get_semester_type_display()}"
                semester_map[key] = sem

            # 考核类型映射
            term_type_map = {tt.name: tt for tt in TermType.objects.all()}

            # 部门映射
            depart_map = {ad.name: ad for ad in AssessDepart.objects.all()}

            # 教师映射 (姓名->对象)
            teacher_map = {t.name: t for t in UserInfo.objects.all()}

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 跳过空行
                    if not any(row):
                        continue

                    # 解析学期 (格式: "2023-2024上学期")
                    semester_str = row[0]
                    if not semester_str:
                        raise ValueError("学期不能为空")

                    if semester_str not in semester_map:
                        # 尝试创建新学期
                        year = semester_str[:-3]  # 去掉后3个字符(学期名)
                        semester_type = 'last' if '上' in semester_str else 'next'
                        sem, created_sem = Semester.objects.get_or_create(
                            year=year,
                            semester_type=semester_type
                        )
                        semester_map[semester_str] = sem

                    # 获取其他关联对象
                    term_type_name = row[1]
                    if not term_type_name:
                        raise ValueError("考核类型不能为空")

                    term_type = term_type_map.get(term_type_name)
                    if not term_type:
                        term_type = TermType.objects.create(
                            name=term_type_name)
                        term_type_map[term_type_name] = term_type

                    depart_name = row[3]
                    if not depart_name:
                        raise ValueError("考核部门不能为空")

                    assess_depart = depart_map.get(depart_name)
                    if not assess_depart:
                        assess_depart = AssessDepart.objects.create(
                            name=depart_name)
                        depart_map[depart_name] = assess_depart

                    teacher_name = row[4]
                    if not teacher_name:
                        raise ValueError("教师姓名不能为空")

                    teacher = teacher_map.get(teacher_name)
                    if not teacher:
                        raise ValueError(f"教师 '{teacher_name}' 不存在")

                    # 新增：处理班级号
                    class_number = row[5]
                    if class_number is None:
                        raise ValueError("班级号不能为空")
                    try:
                        class_number = int(class_number)
                    except (TypeError, ValueError):
                        raise ValueError(f"班级号必须是整数: {class_number}")

                    # 转换字段值
                    def safe_float(value, default=0.0):
                        try:
                            return round(value, 3) if value is not None else default
                        except (TypeError, ValueError):
                            return default

                    defaults = {
                        'teacher': teacher,
                        'assess_time': row[2] or datetime.date.today().isoformat(),
                        'assess_depart': assess_depart,
                        'class_number': class_number,
                        'manage_score': safe_float(row[6]),
                        'safety_score': safe_float(row[7]),
                        'class_score': safe_float(row[8]),
                        'remark': row[9] or '',
                    }
                    
                    # 关键修改：使用模型的唯一约束字段作为查询条件
                    obj, created = HeadTeacherFinalAssess.objects.get_or_create(
                        semester=semester_map[semester_str],
                        term_type=term_type,
                        class_number=class_number,
                        assess_depart=assess_depart,
                        defaults=defaults
                    )
                    
                    # 如果是更新操作，需要先更新字段再保存
                    if not created:
                        # 只更新教师和其他可更新字段
                        obj.teacher = teacher
                        obj.assess_time = defaults['assess_time']
                        obj.assess_depart = defaults['assess_depart']
                        obj.manage_score = defaults['manage_score']
                        obj.safety_score = defaults['safety_score']
                        obj.class_score = defaults['class_score']
                        obj.remark = defaults['remark']

                    # 触发save方法以计算total_score
                    obj.save()

                    success_count += 1
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                except Exception as e:
                    errors.append(f"第 {row_num} 行错误: {str(e)}")

            if errors:
                messages.warning(
                    request, f"成功导入 {success_count} 条（新增:{created_count}, 更新:{updated_count}），失败 {len(errors)} 条")
                for error in errors[:5]:  # 最多显示5条错误
                    messages.error(request, error)
                if len(errors) > 5:
                    messages.info(request, f"还有 {len(errors)-5} 条错误未显示...")
            else:
                messages.success(
                    request, f"成功导入 {success_count} 条数据（新增:{created_count}, 更新:{updated_count}）")

        except Exception as e:
            messages.error(request, f"文件处理错误: {str(e)}")

    return redirect('assessments:headteacher_end_list')


def headteacher_end_export(request):
    """导出班主任考核数据"""
    # 获取筛选参数
    semester_id = request.GET.get('semester')
    term_type_id = request.GET.get('term_type')
    assess_depart_id = request.GET.get('assess_depart')
    teacher_name = request.GET.get('teacher_name')
    subject_id = request.GET.get('subject')

    # 构建查询条件
    query = Q()
    if semester_id and semester_id != 'all':
        query &= Q(semester_id=semester_id)
    if term_type_id and term_type_id != 'all':
        query &= Q(term_type_id=term_type_id)
    if assess_depart_id and assess_depart_id != 'all':
        query &= Q(assess_depart_id=assess_depart_id)
    if teacher_name:
        query &= Q(teacher__name__icontains=teacher_name)
    if subject_id and subject_id != 'all':
        query &= Q(teacher__subject_id=subject_id)

    # 获取数据
    queryset =HeadTeacherFinalAssess.objects.filter(query).select_related(
        'semester', 'term_type', 'assess_depart', 'teacher', 'teacher__subject'
    ).order_by('id')

    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "班主任考核数据"

    # 设置表头
    headers = [
        '序号', '学期', '考核类型', '考核时间', '考核部门', '班主任',
        '班级', '常规管理成绩', '闭环式安全成绩', '班级成绩',
        '总成绩', '名次', '备注',
    ]

    # 添加表头行
    ws.append(headers)

    # 设置表头样式
    header_font = Font(bold=True, size=10)
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # 添加数据行
    for row_num, obj in enumerate(queryset, start=2):
        data = [
            row_num - 1,
            obj.semester.__str__(),
            obj.term_type.name,
            obj.assess_time,
            obj.assess_depart.name,
            obj.teacher.name,
            obj.class_number,
            obj.manage_score,
            obj.safety_score,    
            obj.class_score,    
            obj.total_score,
            obj.rank,
            obj.remark,
        ]

        ws.append(data)

        # 设置数据行样式
        for cell in ws[row_num]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 设置文件名
    filename = f"教师期中考核数据_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

    # 准备响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # 保存工作簿到响应
    wb.save(response)

    return response


@superuser_required
def headteacher_end_update_rank(request):
    """更新班主任期末考核记录的名次，并将记录状态改为“已公示”
    
    核心流程：
    1. 接收POST请求上传的期末考核名次Excel文件（含考核部门字段）
    2. 解析Excel数据（学期、考核类型、考核部门、教师、班级、名次）
    3. 多维度校验数据（含部门存在性、“部门+班级”唯一对应班主任）
    4. 匹配并更新HeadTeacherFinalAssess模型的名次与公示状态
    5. 收集处理结果并反馈前端，最终跳转回期末考核列表
    """
    
    if request.method == "POST":
        # 1. 接收并校验Excel文件是否上传
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "请选择需要上传的期末考核名次Excel文件")
            return redirect('assessments:headteacher_end_list')  # 跳转回期末考核列表页

        try:
            # 2. 初始化Excel处理工具与统计变量
            wb = load_workbook(excel_file)  # 加载Excel文件（支持.xlsx格式）
            ws = wb.active  # 获取Excel活动工作表（默认第一个工作表）
            
            # 统计与错误收集变量
            errors = []  # 存储每一行数据的处理错误信息
            success_count = 0  # 成功更新的期末考核记录数
            updated_count = 0  # 实际修改的记录数
            not_found_count = 0  # 未找到匹配期末考核记录的数量

            # 3. 预加载数据库关联数据，生成“名称-对象”映射
            # 3.1 学期映射：key为“学年+学期类型”（如“2023-2024第二学期”）
            semester_map = {}
            for sem in Semester.objects.all():
                key = f"{sem.year}{sem.get_semester_type_display()}"
                semester_map[key] = sem

            # 3.2 考核类型映射：key为考核类型名称（如“期末考核”）
            term_type_map = {tt.name: tt for tt in TermType.objects.all()}

            # 3.3 教师映射：key为教师姓名
            teacher_map = {t.name: t for t in UserInfo.objects.all()}

            # 3.4 考核部门映射：key为部门名称（如“高一年级组”）
            assess_depart_map = {ad.name: ad for ad in AssessDepart.objects.all()}
            
            # 3.5 【核心】班级-部门-班主任映射：key为“部门ID_班级号”（解决同班级号跨部门问题）
            class_depart_teacher_map = {
                f"{cls.assess_depart.id}_{cls.class_number}": cls.teacher 
                for cls in HeadTeacherFinalAssess.objects.all()
            }

            # 4. 遍历Excel行数据（从第2行开始，跳过表头）
            # Excel列顺序：0=学期, 1=考核类型, 2=考核部门, 3=教师姓名, 4=班级号, 5=名次
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 4.1 跳过空行
                    if not any(row):
                        continue

                    # 4.2 提取Excel当前行关键字段（row[2]为考核部门）
                    semester_str = row[0]          # 列1：学期
                    term_type_name = row[1]        # 列2：考核类型
                    assess_depart_name = row[2]    # 列3：考核部门
                    teacher_name = row[3]          # 列4：教师姓名
                    class_number = row[4]          # 列5：班级号
                    rank_value = row[5]            # 列6：名次

                    # 4.3 字段预处理：空值处理+字符串去空格
                    semester_str = str(semester_str).strip() if semester_str else None
                    term_type_name = str(term_type_name).strip() if term_type_name else None
                    assess_depart_name = str(assess_depart_name).strip() if assess_depart_name else None
                    teacher_name = str(teacher_name).strip() if teacher_name else None

                    # 4.4 基础校验：必填字段不能为空（包含考核部门）
                    required_fields = [semester_str, term_type_name, assess_depart_name, 
                                      teacher_name, class_number, rank_value]
                    if not all(required_fields):
                        raise ValueError("学期、考核类型、考核部门、教师姓名、班级号、名次均为必填字段，不可为空")

                    # 4.5 格式校验：班级号、名次、部门
                    # 班级号转整数
                    try:
                        class_number = int(class_number)
                    except (TypeError, ValueError):
                        raise ValueError(f"班级号格式错误，必须为整数（当前值：{class_number}）")
                    
                    # 名次格式校验
                    try:
                        rank_float = float(rank_value)
                        if rank_float <= 0:
                            raise ValueError("名次必须为大于0的数字（如1、2、3.5）")
                        # 若需整数名次可启用：rank_float = int(rank_float)
                    except (TypeError, ValueError):
                        raise ValueError(f"名次格式错误，必须为有效数字（当前值：{rank_value}）")
                    
                    # 考核部门存在性校验
                    assess_depart = assess_depart_map.get(assess_depart_name)
                    if not assess_depart:
                        raise ValueError(f"考核部门「{assess_depart_name}」不存在，请确认部门名称与系统配置一致")

                    # 4.6 关联对象存在性校验
                    # 校验学期
                    semester = semester_map.get(semester_str)
                    if not semester:
                        raise ValueError(f"学期不存在（当前值：{semester_str}），请确认格式为“学年+学期类型”")
                    
                    # 校验考核类型
                    term_type = term_type_map.get(term_type_name)
                    if not term_type:
                        raise ValueError(f"考核类型不存在（当前值：{term_type_name}），请确认类型名称正确")
                    
                    # 校验教师
                    teacher = teacher_map.get(teacher_name)
                    if not teacher:
                        raise ValueError(f"教师不存在（当前值：{teacher_name}），请确认教师姓名正确")

                    # 4.7 【核心校验】班主任与部门+班级的唯一性匹配
                    check_key = f"{assess_depart.id}_{class_number}"
                    if class_depart_teacher_map.get(check_key) != teacher:
                        raise ValueError(
                            f"教师「{teacher_name}」不是「{assess_depart_name}」部门下班级「{class_number}」的班主任，无法更新考核名次"
                        )

                    # 4.8 匹配并更新期末考核记录（包含考核部门条件）
                    try:
                        # 严格匹配模型的unique_together约束：semester+term_type+assess_depart+class_number
                        assess = HeadTeacherFinalAssess.objects.get(
                            semester=semester,
                            term_type=term_type,
                            assess_depart=assess_depart,  # 关键：加入考核部门条件
                            class_number=class_number,
                            teacher=teacher
                        )

                        # 更新名次和公示状态
                        assess.rank = rank_float
                        assess.is_published = True
                        assess.save()
                        
                        success_count += 1
                        updated_count += 1

                    except HeadTeacherFinalAssess.DoesNotExist:
                        not_found_count += 1
                        errors.append(
                            f"第 {row_num} 行: 未找到匹配的期末考核记录 "
                            f"[部门：{assess_depart_name}，教师：{teacher_name}，班级：{class_number}，学期：{semester_str}]"
                        )

                except Exception as e:
                    errors.append(f"第 {row_num} 行错误：{str(e)}")

            # 5. 处理结果反馈
            result_message = f"成功更新 {success_count} 条期末考核记录"
            if not_found_count > 0:
                result_message += f"，未找到 {not_found_count} 条匹配记录"
            
            if errors:
                messages.warning(request, result_message)
                for error in errors[:5]:
                    messages.error(request, error)
                if len(errors) > 5:
                    messages.info(request, f"还有 {len(errors)-5} 条错误未显示，可检查Excel对应行数据")
            else:
                messages.success(request, result_message)

        except Exception as e:
            messages.error(request, f"文件处理失败：{str(e)}（建议检查Excel文件格式是否为.xlsx）")

    # 跳转回期末考核列表页
    return redirect('assessments:headteacher_end_list')