from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from django.db.models import Q
from django.db import transaction

from utils.pagination import Pagination
from utils.bootstrap import BootStrapModelForm

from assessments.models import HeadTeacherSemester, AssessDepart, Semester, TermType, HeadTeacherMidAssess, HeadTeacherFinalAssess
from accounts.models import UserInfo, Subject
from utils.user_decorator import superuser_required


class AssessModelForm(BootStrapModelForm):
    class Meta:
        model = HeadTeacherSemester
        # 字段，所有字段
        fields = '__all__'


def headteacher_term_list(request):
    # 获取所有可选数据
    semesters = Semester.objects.order_by('-id')
    term_types = TermType.objects.all()
    assess_departs = AssessDepart.objects.all()
    subjects = Subject.objects.all()

    # 初始化查询条件
    semester_id = request.GET.get('semester')
    term_type_id = request.GET.get('term_type')
    assess_depart_id = request.GET.get('assess_depart')
    teacher_name = request.GET.get('teacher_name')
    subject_id = request.GET.get('subject')  # 新增学科参数

    # 构建查询条件
    query = Q()
    if semester_id and semester_id != 'all':
        query &= Q(semester_id=semester_id)
    if term_type_id and term_type_id != 'all':
        query &= Q(term_type_id=term_type_id)
    if assess_depart_id and assess_depart_id != 'all':
        query &= Q(assess_depart_id=assess_depart_id)
    if teacher_name:
        query &= Q(teacher__name__icontains=teacher_name)
    if subject_id and subject_id != 'all':
        query &= Q(teacher__subject_id=subject_id)
        
    # 权限控制：普通用户只能查看自己的记录
    if not (request.user.is_superuser or request.user.groups.filter(name='管理员').exists()):
        query &= Q(teacher=request.user)

    # 应用查询条件
    queryset = HeadTeacherSemester.objects.filter(
        query).order_by('id')

    page_object = Pagination(request, queryset)

    content = {
        'queryset': page_object.page_queryset,
        'page_string': page_object.html(),
        'semesters': semesters,
        'term_types': term_types,
        'assess_departs': assess_departs,
        'subjects': subjects,
        'selected_semester': semester_id if semester_id else 'all',
        'selected_term_type': term_type_id if term_type_id else 'all',
        'selected_assess_depart': assess_depart_id if assess_depart_id else 'all',
        'teacher_name': teacher_name if teacher_name else '',
        'selected_subject': subject_id if subject_id else 'all',

    }
    return render(request, 'headteacher_term_list.html', content)


@superuser_required
def headteacher_term_delete(request):
    """删除"""
    nid = request.GET.get('nid')
    HeadTeacherSemester.objects.filter(id=nid).delete()
    return redirect('assessments:headteacher_term_list')


@superuser_required
def headteacher_term_edit(request, pk):
    instance = get_object_or_404(HeadTeacherSemester, pk=pk)
    form = AssessModelForm(request.POST or None, instance=instance)

    if request.method == 'POST':
        if form.is_valid():
            # 1. 从表单/实例中获取班级号等参数（编辑时可能修改班级，优先从表单取）
            teacher = form.cleaned_data['teacher']
            semester = form.cleaned_data['semester']
            class_number = form.cleaned_data['class_number']  # 从表单提取班级号
            assess_depart = form.cleaned_data['assess_depart']
            term_type = form.cleaned_data['term_type']

            # 2. 精准查询期中成绩（匹配唯一约束）
            mid_assess = HeadTeacherMidAssess.objects.filter(
                semester=semester,
                # term_type=term_type,
                assess_depart=assess_depart,
                class_number=class_number  # 核心：班级号匹配
            ).first()
            
            # 3. 精准查询期末成绩
            final_assess = HeadTeacherFinalAssess.objects.filter(
                semester=semester,
                # term_type=term_type,
                assess_depart=assess_depart,
                class_number=class_number  # 核心：班级号匹配
            ).first()
            
            # 4. 保存关联和总成绩
            instance.mid_score = mid_assess
            instance.final_score = final_assess
            instance.total_score = (mid_assess.total_score if mid_assess else 0) + \
                                 (final_assess.total_score if final_assess else 0)
            instance.save()
            return redirect('assessments:headteacher_term_list')

    # 上下文逻辑不变
    context = {
        'form': form,
        'title': '考核记录',
        'instance': instance,
        'show_class_field':True,
        'show_mid_score': True,
        'show_final_score': True
    }
    return render(request, 'assess_change.html', context)


@superuser_required
def headteacher_term_add(request):
    form = AssessModelForm()
    teachers = UserInfo.objects.all()
    if request.method == 'POST':
        form = AssessModelForm(request.POST)
        if form.is_valid():
            # 1. 从表单中获取班级号（关键：新增班级参数）
            teacher = form.cleaned_data['teacher']
            semester = form.cleaned_data['semester']
            class_number = form.cleaned_data['class_number']  # 从表单提取班级号
            assess_depart = form.cleaned_data['assess_depart']  # 唯一约束包含考核部门，需一并传递
            term_type = form.cleaned_data['term_type']  # 唯一约束包含考核类型，需一并传递

            # 2. 查询期中成绩：补充班级号、考核部门、考核类型（匹配唯一约束）
            mid_assess = HeadTeacherMidAssess.objects.filter(
                semester=semester,
                # term_type=term_type,
                assess_depart=assess_depart,
                class_number=class_number  # 核心：通过班级号精准匹配
            ).first()
            
            # 3. 查询期末成绩：同样补充班级号等参数
            final_assess = HeadTeacherFinalAssess.objects.filter(
                semester=semester,
                # term_type=term_type,
                assess_depart=assess_depart,
                class_number=class_number  # 核心：通过班级号精准匹配
            ).first()
            
            # 4. 后续保存逻辑不变
            instance = form.save(commit=False)
            instance.mid_score = mid_assess
            instance.final_score = final_assess
            instance.total_score = (mid_assess.total_score if mid_assess else 0) + \
                                 (final_assess.total_score if final_assess else 0)
            instance.save()
            
            return redirect('assessments:headteacher_term_list')
    
    return render(request, 'assess_change.html', {'form': form, 'title': '新建考核','show_mid_score': True, 'show_final_score': True, 'show_class_field':True,})


# 下面是批量导入需要的功能
@superuser_required
@transaction.atomic
def headteacher_term_import(request):
    """批量导入班主任学期考核成绩并自动关联期中期末成绩，包含班级信息匹配"""
    if request.method == "POST":
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "请选择Excel文件")
            return redirect('assessments:headteacher_term_list')

        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            errors = []
            success_count = 0
            created_count = 0
            updated_count = 0

            # 预加载基础数据映射（优化查询性能）
            semester_map = {f"{sem.year}{sem.get_semester_type_display()}": sem 
                           for sem in Semester.objects.all()}
            term_type_map = {tt.name: tt for tt in TermType.objects.all()}
            depart_map = {ad.name: ad for ad in AssessDepart.objects.all()}
            teacher_map = {t.name: t for t in UserInfo.objects.all()}
            
            # 预加载期中期末成绩映射（使用唯一约束字段作为键）
            # 键结构：(学期ID, 考核类型ID, 考核部门ID, 班级号)
            mid_assess_map = {}
            for mid in HeadTeacherMidAssess.objects.all():
                key = (mid.semester_id, mid.term_type_id, mid.assess_depart_id, mid.class_number)
                mid_assess_map[key] = mid
                
            final_assess_map = {}
            for final in HeadTeacherFinalAssess.objects.all():
                key = (final.semester_id, final.term_type_id, final.assess_depart_id, final.class_number)
                final_assess_map[key] = final

            # 遍历Excel行（从第2行开始，跳过表头）
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 跳过空行
                    if not any(row):
                        continue

                    # 解析Excel字段（根据实际Excel列顺序调整索引）
                    # 假设Excel列顺序：学期、考核类型、考核时间、部门、教师姓名、班级号、备注
                    semester_str = str(row[0]).strip() if row[0] else ""
                    term_type_name = str(row[1]).strip() if row[1] else ""
                    assess_time = row[2] if row[2] else datetime.date.today().isoformat()
                    depart_name = str(row[3]).strip() if row[3] else ""
                    teacher_name = str(row[4]).strip() if row[4] else ""
                    class_number = row[5] if (len(row) > 5 and row[5] is not None) else None
                    remark = str(row[6]).strip() if (len(row) > 6 and row[6] is not None) else ""

                    # 班级号验证与转换
                    if class_number is None:
                        raise ValueError("班级号不能为空")
                    try:
                        class_number = int(class_number)
                        if class_number <= 0:
                            raise ValueError("班级号必须为正整数")
                    except (TypeError, ValueError):
                        raise ValueError(f"班级号格式错误: {class_number}（应为正整数）")

                    # 基础字段校验
                    if not semester_str:
                        raise ValueError("学期不能为空")
                    if not term_type_name:
                        raise ValueError("考核类型不能为空")
                    if not depart_name:
                        raise ValueError("考核部门不能为空")
                    if not teacher_name:
                        raise ValueError("教师姓名不能为空")

                    # 处理学期（自动创建不存在的学期）
                    if semester_str not in semester_map:
                        try:
                            # 解析学期格式（例如"2023-2024上学期"）
                            if '上学期' in semester_str:
                                year = semester_str.replace('上学期', '').strip()
                                semester_type = 'last'
                            elif '下学期' in semester_str:
                                year = semester_str.replace('下学期', '').strip()
                                semester_type = 'next'
                            else:
                                raise ValueError(f"学期格式错误: {semester_str}（应为'XXXX-XXXX上学期'或'XXXX-XXXX下学期'）")
                            
                            sem, created = Semester.objects.get_or_create(
                                year=year,
                                semester_type=semester_type
                            )
                            semester_map[semester_str] = sem
                        except Exception as e:
                            raise ValueError(f"处理学期失败: {str(e)}")
                    semester = semester_map[semester_str]

                    # 处理考核类型（自动创建不存在的类型）
                    term_type = term_type_map.get(term_type_name)
                    if not term_type:
                        term_type = TermType.objects.create(name=term_type_name)
                        term_type_map[term_type_name] = term_type

                    # 处理考核部门（自动创建不存在的部门）
                    assess_depart = depart_map.get(depart_name)
                    if not assess_depart:
                        assess_depart = AssessDepart.objects.create(name=depart_name)
                        depart_map[depart_name] = assess_depart

                    # 处理教师（必须已存在）
                    teacher = teacher_map.get(teacher_name)
                    if not teacher:
                        raise ValueError(f"教师 '{teacher_name}' 不存在于系统中")

                    # 构建查询键（匹配唯一约束的4个字段）
                    query_key = (semester.id, term_type.id, assess_depart.id, class_number)
                    
                    # 查找期中成绩
                    mid_assess = mid_assess_map.get(query_key)
                    if not mid_assess:
                        # 二次查询避免映射缓存遗漏(term_type在学期新建、编辑、批量创建的时候不用查询，因为学期、期中、期末不一样)
                        mid_assess = HeadTeacherMidAssess.objects.filter(
                            semester_id=semester.id,
                            # term_type_id=term_type.id,
                            assess_depart_id=assess_depart.id,
                            class_number=class_number
                        ).first()
                        if mid_assess:
                            mid_assess_map[query_key] = mid_assess  # 更新缓存
                    
                    # 查找期末成绩
                    final_assess = final_assess_map.get(query_key)
                    if not final_assess:
                        final_assess = HeadTeacherFinalAssess.objects.filter(
                            semester_id=semester.id,
                            # term_type_id=term_type.id,
                            assess_depart_id=assess_depart.id,
                            class_number=class_number
                        ).first()
                        if final_assess:
                            final_assess_map[query_key] = final_assess  # 更新缓存

                    # 提示未找到关联成绩（非错误，仅提醒）
                    if not mid_assess:
                        messages.warning(request, f"第{row_num}行：未找到对应班级的期中成绩")
                    if not final_assess:
                        messages.warning(request, f"第{row_num}行：未找到对应班级的期末成绩")
                    
                    # 创建或更新学期总评记录
                    defaults = {
                        'assess_time': assess_time,
                        'assess_depart': assess_depart,
                        'teacher': teacher,
                        'mid_score': mid_assess,
                        'final_score': final_assess,
                        'remark': remark,
                        # 自动计算总成绩
                        'total_score': round(
                            (mid_assess.total_score if mid_assess else 0) + 
                            (final_assess.total_score if final_assess else 0), 
                            3
                        )
                    }
                    
                    # 根据唯一约束条件查询记录
                    obj, created = HeadTeacherSemester.objects.get_or_create(
                        semester=semester,
                        term_type=term_type,
                        assess_depart=assess_depart,
                        class_number=class_number,
                        defaults=defaults
                    )
                    
                    # 如果是更新操作，同步字段值
                    if not created:
                        for key, value in defaults.items():
                            setattr(obj, key, value)
                        obj.save()
                        updated_count += 1
                    else:
                        created_count += 1
                    
                    success_count += 1

                except Exception as e:
                    errors.append(f"第 {row_num} 行错误: {str(e)}")

            # 处理导入结果反馈
            if errors:
                msg = f"成功导入 {success_count} 条（新增:{created_count}, 更新:{updated_count}），失败 {len(errors)} 条"
                messages.warning(request, msg)
                for error in errors[:5]:  # 最多显示5条错误
                    messages.error(request, error)
                if len(errors) > 5:
                    messages.info(request, f"还有 {len(errors)-5} 条错误未显示...")
            else:
                messages.success(
                    request, f"成功导入 {success_count} 条数据（新增:{created_count}, 更新:{updated_count}）")

        except Exception as e:
            messages.error(request, f"文件处理错误: {str(e)}")

    return redirect('assessments:headteacher_term_list')



def headteacher_term_export(request):
    """导出教师学期考核数据"""
    # 获取筛选参数
    semester_id = request.GET.get('semester')
    term_type_id = request.GET.get('term_type')
    assess_depart_id = request.GET.get('assess_depart')
    teacher_name = request.GET.get('teacher_name')
    subject_id = request.GET.get('subject')

    # 构建查询条件
    query = Q()
    if semester_id and semester_id != 'all':
        query &= Q(semester_id=semester_id)
    if term_type_id and term_type_id != 'all':
        query &= Q(term_type_id=term_type_id)
    if assess_depart_id and assess_depart_id != 'all':
        query &= Q(assess_depart_id=assess_depart_id)
    if teacher_name:
        query &= Q(teacher__name__icontains=teacher_name)
    if subject_id and subject_id != 'all':
        query &= Q(teacher__subject_id=subject_id)

    # 获取数据
    queryset = HeadTeacherSemester.objects.filter(query).select_related(
        'semester', 'term_type', 'assess_depart', 'teacher', 'teacher__subject'
    ).order_by('id')

    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "教师学期考核数据"

    # 设置表头
    headers = [
        '序号', '学期', '考核类型', '考核时间', '考核部门', '班主任','班级', 
        '期中成绩', '期末成绩', '学期成绩', '名次', '备注',
    ]

    # 添加表头行
    ws.append(headers)

    # 设置表头样式
    header_font = Font(bold=True, size=10)
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # 添加数据行
    for row_num, obj in enumerate(queryset, start=2):
        data = [
            row_num - 1,
            obj.semester.__str__(),
            obj.term_type.name,
            obj.assess_time,
            obj.assess_depart.name,
            obj.teacher.name,
            obj.class_number,
            obj.mid_score.total_score if obj.mid_score else '',
            obj.final_score.total_score if obj.final_score else '',
            obj.total_score,
            obj.rank,
            obj.remark,
        ]

        ws.append(data)

        # 设置数据行样式
        for cell in ws[row_num]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 设置文件名
    filename = f"教师学期考核数据_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

    # 准备响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # 保存工作簿到响应
    wb.save(response)

    return response


@superuser_required
def headteacher_term_update_rank(request):
    """更新班主任学期考核记录的名次，并将记录状态改为“已公示”
    
    核心流程：
    1. 接收POST请求上传的学期考核名次Excel文件（含考核部门字段，第3列）
    2. 解析Excel数据（学期、考核类型、考核部门、教师、班级、名次）
    3. 多维度校验数据（部门存在性、“部门+班级”唯一对应班主任等）
    4. 匹配并更新HeadTeacherSemester模型的名次与公示状态
    5. 收集处理结果并反馈前端，最终跳转回学期考核列表
    """
    
    if request.method == "POST":
        # 1. 接收并校验Excel文件是否上传
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "请选择需要上传的学期考核名次Excel文件")
            return redirect('assessments:headteacher_term_list')  # 跳转回学期考核列表页

        try:
            # 2. 初始化Excel处理工具与统计变量
            wb = load_workbook(excel_file)  # 加载Excel文件（支持.xlsx格式，需确保文件未损坏）
            ws = wb.active  # 获取Excel活动工作表（默认第一个工作表，需与表头对应）
            
            # 统计与错误收集变量
            errors = []  # 存储每一行的处理错误信息（含行号，方便定位）
            success_count = 0  # 成功更新的学期考核记录数
            updated_count = 0  # 实际修改的记录数（与success_count一致，预留扩展）
            not_found_count = 0  # 未找到匹配记录的数量

            # 3. 预加载数据库关联数据，生成“名称-对象”映射（减少循环内查询，提升性能）
            # 3.1 学期映射：key=“学年+学期类型”（如“2023-2024第一学期”）
            semester_map = {}
            for sem in Semester.objects.all():
                key = f"{sem.year}{sem.get_semester_type_display()}"
                semester_map[key] = sem

            # 3.2 考核类型映射：key=考核类型名称（如“学期综合考核”）
            term_type_map = {tt.name: tt for tt in TermType.objects.all()}

            # 3.3 教师映射：key=教师姓名（如“张三”）
            teacher_map = {t.name: t for t in UserInfo.objects.all()}

            # 3.4 考核部门映射：key=部门名称（如“高一年级组”），新增！匹配Excel第3列
            assess_depart_map = {ad.name: ad for ad in AssessDepart.objects.all()}
            
            # 3.5 【核心修正】班级-部门-班主任映射：key=“部门ID_班级号”（如“1_202301”）
            # 解决“不同部门同班级号”的班主任匹配问题，符合模型unique_together约束
            class_depart_teacher_map = {
                f"{cls.assess_depart.id}_{cls.class_number}": cls.teacher 
                for cls in HeadTeacherSemester.objects.all()
            }

            # 4. 遍历Excel行数据（从第2行开始，跳过表头；行号从2开始，与Excel实际行号对齐）
            # Excel列顺序：0=学期, 1=考核类型, 2=考核部门（新增）, 3=教师姓名, 4=班级号, 5=名次
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 4.1 跳过空行：若当前行所有单元格均为空，直接跳过
                    if not any(row):
                        continue

                    # 4.2 提取Excel字段（严格对应列顺序，row[2]为考核部门）
                    semester_str = row[0]          # 列1：学期
                    term_type_name = row[1]        # 列2：考核类型
                    assess_depart_name = row[2]    # 列3：考核部门（新增，核心字段）
                    teacher_name = row[3]          # 列4：教师姓名
                    class_number = row[4]          # 列5：班级号
                    rank_value = row[5]            # 列6：名次

                    # 4.3 字段预处理：空值处理+去空格（避免“名称前后空格”导致的匹配失败）
                    semester_str = str(semester_str).strip() if semester_str else None
                    term_type_name = str(term_type_name).strip() if term_type_name else None
                    assess_depart_name = str(assess_depart_name).strip() if assess_depart_name else None
                    teacher_name = str(teacher_name).strip() if teacher_name else None

                    # 4.4 基础校验：必填字段不能为空（新增考核部门）
                    required_fields = [semester_str, term_type_name, assess_depart_name, 
                                      teacher_name, class_number, rank_value]
                    if not all(required_fields):
                        raise ValueError("学期、考核类型、考核部门、教师姓名、班级号、名次均为必填字段，不可为空")

                    # 4.5 格式校验：班级号、名次、考核部门
                    # 班级号转整数（匹配模型IntegerField类型）
                    try:
                        class_number = int(class_number)
                    except (TypeError, ValueError):
                        raise ValueError(f"班级号格式错误，必须为整数（当前值：{class_number}）")
                    
                    # 名次校验：必须为大于0的有效数字（支持小数，需整数可取消注释）
                    try:
                        rank_float = float(rank_value)
                        if rank_float <= 0:
                            raise ValueError("名次必须为大于0的数字（如1、2、3.5）")
                        # 若业务要求名次为整数，启用：rank_float = int(rank_float)
                    except (TypeError, ValueError):
                        raise ValueError(f"名次格式错误，必须为有效数字（当前值：{rank_value}）")
                    
                    # 考核部门存在性校验（新增！确保部门在系统中已配置）
                    assess_depart = assess_depart_map.get(assess_depart_name)
                    if not assess_depart:
                        raise ValueError(f"考核部门「{assess_depart_name}」不存在，请确认部门名称与系统配置一致")

                    # 4.6 关联对象存在性校验：学期、考核类型、教师
                    # 校验学期
                    semester = semester_map.get(semester_str)
                    if not semester:
                        raise ValueError(f"学期不存在（当前值：{semester_str}），请确认格式为“学年+学期类型”（如2023-2024第一学期）")
                    
                    # 校验考核类型
                    term_type = term_type_map.get(term_type_name)
                    if not term_type:
                        raise ValueError(f"考核类型不存在（当前值：{term_type_name}），请确认类型名称正确（如“学期综合考核”）")
                    
                    # 校验教师
                    teacher = teacher_map.get(teacher_name)
                    if not teacher:
                        raise ValueError(f"教师不存在（当前值：{teacher_name}），请确认教师姓名正确（区分大小写/空格）")

                    # 4.7 【核心校验】班主任与“部门+班级”的唯一性匹配
                    # 生成唯一key：部门ID_班级号（确保跨部门同班级号能正确区分）
                    check_key = f"{assess_depart.id}_{class_number}"
                    if class_depart_teacher_map.get(check_key) != teacher:
                        raise ValueError(
                            f"教师「{teacher_name}」不是「{assess_depart_name}」部门下班级「{class_number}」的班主任，无法更新学期考核名次"
                        )

                    # 4.8 匹配并更新学期考核记录（查询条件必须包含assess_depart）
                    try:
                        # 严格匹配模型unique_together约束：semester+term_type+assess_depart+class_number
                        assess = HeadTeacherSemester.objects.get(
                            semester=semester,
                            term_type=term_type,
                            assess_depart=assess_depart,  # 新增！关键查询条件
                            class_number=class_number,
                            teacher=teacher  # 冗余校验，确保教师匹配
                        )

                        # 更新名次和公示状态
                        assess.rank = rank_float
                        assess.is_published = True
                        assess.save()  # 触发模型save方法，自动计算总成绩

                        # 更新统计计数
                        success_count += 1
                        updated_count += 1

                    # 处理“未找到匹配记录”的情况
                    except HeadTeacherSemester.DoesNotExist:
                        not_found_count += 1
                        errors.append(
                            f"第 {row_num} 行: 未找到匹配的学期考核记录 "
                            f"[部门：{assess_depart_name}，教师：{teacher_name}，班级：{class_number}，学期：{semester_str}]"
                        )

                # 捕获当前行的所有错误，不中断整体循环
                except Exception as e:
                    errors.append(f"第 {row_num} 行错误：{str(e)}")

            # 5. 处理结果反馈：拼接信息并传递给前端
            result_message = f"成功更新 {success_count} 条学期考核记录"
            if not_found_count > 0:
                result_message += f"，未找到 {not_found_count} 条匹配记录"
            
            # 有错误时显示警告+前5条错误，无错误显示成功信息
            if errors:
                messages.warning(request, result_message)
                for error in errors[:5]:
                    messages.error(request, error)
                if len(errors) > 5:
                    messages.info(request, f"还有 {len(errors)-5} 条错误未显示，可按行号检查Excel对应数据")
            else:
                messages.success(request, result_message)

        # 捕获Excel文件处理的全局错误（如文件损坏、格式不支持等）
        except Exception as e:
            messages.error(request, f"文件处理失败：{str(e)}（建议检查Excel文件格式是否为.xlsx，或文件是否被占用）")

    # 无论处理结果如何，均跳转回学期考核列表页
    return redirect('assessments:headteacher_term_list')