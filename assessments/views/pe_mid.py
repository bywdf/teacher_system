from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from django.db.models import Q
from django.db import transaction

from utils.pagination import Pagination
from utils.bootstrap import BootStrapModelForm

from assessments.models import PeTeacherMidAssess, AssessDepart, Semester, TermType
from accounts.models import UserInfo, Subject
from utils.user_decorator import superuser_required


class MidAssessModelForm(BootStrapModelForm):
    class Meta:
        model = PeTeacherMidAssess
        # 字段，所有字段
        fields = '__all__'


def pe_mid_list(request):
    # 获取所有可选数据
    semesters = Semester.objects.order_by('-id')
    term_types = TermType.objects.all()
    assess_departs = AssessDepart.objects.all()
    subjects = Subject.objects.all()

    # 初始化查询条件
    semester_id = request.GET.get('semester')
    term_type_id = request.GET.get('term_type')
    assess_depart_id = request.GET.get('assess_depart')
    teacher_name = request.GET.get('teacher_name')
    subject_id = request.GET.get('subject')  # 新增学科参数

    # 构建查询条件
    query = Q()
    if semester_id and semester_id != 'all':
        query &= Q(semester_id=semester_id)
    if term_type_id and term_type_id != 'all':
        query &= Q(term_type_id=term_type_id)
    if assess_depart_id and assess_depart_id != 'all':
        query &= Q(assess_depart_id=assess_depart_id)
    if teacher_name:
        query &= Q(teacher__name__icontains=teacher_name)
    if subject_id and subject_id != 'all':
        query &= Q(teacher__subject_id=subject_id)
        
    # 权限控制：普通用户只能查看自己的记录
    if not (request.user.is_superuser or request.user.groups.filter(name='管理员').exists()):
        query &= Q(teacher=request.user)

    # 应用查询条件
    queryset = PeTeacherMidAssess.objects.filter(
        query).order_by('id')

    page_object = Pagination(request, queryset)

    content = {
        'queryset': page_object.page_queryset,
        'page_string': page_object.html(),
        'semesters': semesters,
        'term_types': term_types,
        'assess_departs': assess_departs,
        'subjects': subjects,
        'selected_semester': semester_id if semester_id else 'all',
        'selected_term_type': term_type_id if term_type_id else 'all',
        'selected_assess_depart': assess_depart_id if assess_depart_id else 'all',
        'teacher_name': teacher_name if teacher_name else '',
        'selected_subject': subject_id if subject_id else 'all',

    }
    return render(request, 'pe_mid_list.html', content)


@superuser_required
def pe_mid_delete(request):
    """删除"""
    nid = request.GET.get('nid')
    PeTeacherMidAssess.objects.filter(id=nid).delete()
    return redirect('assessments:pe_mid_list')


@superuser_required
def pe_mid_edit(request, pk):
    # 获取要编辑的对象，若不存在则返回404
    instance = get_object_or_404(PeTeacherMidAssess, pk=pk)
    # 创建表单实例，绑定现有数据
    form = MidAssessModelForm(request.POST or None, instance=instance)

    if request.method == 'POST':
        if form.is_valid():
            # 保存更新（可在此处添加额外逻辑，如权限检查、计算字段等）
            form.save()
            return redirect('assessments:pe_mid_list')  # 重定向到列表页
        # 若表单验证失败，保留错误信息并重新渲染页面

    # 渲染编辑页面，传递表单和对象
    context = {
        'form': form,
        'title': '考核记录',
        'instance': instance,
        'show_workload_fields':True,
        'show_major_hours':True,
        'show_teach_book':True,
        'show_kejiancao_hours':True, 
    }
    return render(request, 'assess_change.html', context)


@superuser_required
def pe_mid_add(request):
    """添加"""
    form = MidAssessModelForm()
    # 获取所有教师数据
    teachers = UserInfo.objects.all()
    if request.method == 'POST':
        form = MidAssessModelForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('assessments:pe_mid_list')
        
    content = {
        'form': form,
        'title': '添加考核记录',
        'show_workload_fields':True,
        'show_major_hours':True,
        'show_teach_book':True,
        'show_kejiancao_hours':True,        
    }
    
    return render(request, 'assess_change.html', content)


# 下面是批量导入需要的功能
@superuser_required
@transaction.atomic
def pe_mid_import(request):
    """批量导入考核成绩"""
    if request.method == "POST":
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "请选择Excel文件")
            return redirect('assessments:pe_mid_list')

        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            errors = []
            success_count = 0
            created_count = 0
            updated_count = 0

            # 学期映射字典
            semester_map = {}
            for sem in Semester.objects.all():
                key = f"{sem.year}{sem.get_semester_type_display()}"
                semester_map[key] = sem

            # 考核类型映射
            term_type_map = {tt.name: tt for tt in TermType.objects.all()}

            # 部门映射
            depart_map = {ad.name: ad for ad in AssessDepart.objects.all()}

            # 教师映射 (姓名->对象)
            teacher_map = {t.name: t for t in UserInfo.objects.all()}

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 跳过空行
                    if not any(row):
                        continue

                    # 解析学期 (格式: "2023-2024上学期")
                    semester_str = row[0]
                    if not semester_str:
                        raise ValueError("学期不能为空")

                    # 修改：学期不存在时抛出异常，不再自动创建
                    if semester_str not in semester_map:
                        raise ValueError(f"学期 '{semester_str}' 不存在，请先在系统中创建该学期")

                    # 获取其他关联对象
                    term_type_name = row[1]
                    if not term_type_name:
                        raise ValueError("考核类型不能为空")

                    term_type = term_type_map.get(term_type_name)
                    if not term_type:                      
                        raise ValueError(f"考核类型 '{term_type_name}' 不存在，请先在系统中创建考核类型")

                    depart_name = row[3]
                    if not depart_name:
                        raise ValueError("考核部门不能为空")

                    # 优化后的逻辑：仅允许使用已存在的部门，不存在则报错
                    assess_depart = depart_map.get(depart_name)
                    if not assess_depart:
                        raise ValueError(f"考核部门 '{depart_name}' 不存在，请先在系统中创建该部门")

                    teacher_name = row[4]
                    if not teacher_name:
                        raise ValueError("教师姓名不能为空")

                    teacher = teacher_map.get(teacher_name)
                    if not teacher:
                        raise ValueError(f"教师 '{teacher_name}' 不存在")

                    # 转换字段值
                    def safe_float(value, default=0.0):
                        try:
                            return round(value, 3) if value is not None else default
                        except (TypeError, ValueError):
                            return default

                    defaults={
                            'assess_time': row[2] or datetime.date.today().isoformat(),
                            'assess_depart': assess_depart,
                            'class_hours': safe_float(row[5]),
                            'major_hours': safe_float(row[6]),
                            'kejiancao_hours': safe_float(row[7]),
                            'extra_work_hours': safe_float(row[8]),
                            'total_workload': safe_float(row[9]),
                            'workload_score': safe_float(row[10]),
                            'teach_book': safe_float(row[11]),
                            'remark': row[12] or "",                            
                        }
                    
                    # 创建或更新记录
                    obj, created = PeTeacherMidAssess.objects.get_or_create(
                        teacher=teacher,
                        semester=semester_map[semester_str],
                        term_type=term_type,
                        defaults=defaults
                    )
                    
                    # 如果是更新操作，需要先更新字段再保存
                    if not created:
                        for key, value in defaults.items():
                            setattr(obj, key, value)

                    # 触发save方法以计算total_score
                    obj.save()

                    success_count += 1
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                except Exception as e:
                    errors.append(f"第 {row_num} 行错误: {str(e)}")

            if errors:
                messages.warning(
                    request, f"成功导入 {success_count} 条（新增:{created_count}, 更新:{updated_count}），失败 {len(errors)} 条")
                for error in errors[:5]:  # 最多显示5条错误
                    messages.error(request, error)
                if len(errors) > 5:
                    messages.info(request, f"还有 {len(errors)-5} 条错误未显示...")
            else:
                messages.success(
                    request, f"成功导入 {success_count} 条数据（新增:{created_count}, 更新:{updated_count}）")

        except Exception as e:
            messages.error(request, f"文件处理错误: {str(e)}")

    return redirect('assessments:pe_mid_list')


def pe_mid_export(request):
    """导出教师期中考核数据"""
    # 获取筛选参数
    semester_id = request.GET.get('semester')
    term_type_id = request.GET.get('term_type')
    assess_depart_id = request.GET.get('assess_depart')
    teacher_name = request.GET.get('teacher_name')
    subject_id = request.GET.get('subject')

    # 构建查询条件
    query = Q()
    if semester_id and semester_id != 'all':
        query &= Q(semester_id=semester_id)
    if term_type_id and term_type_id != 'all':
        query &= Q(term_type_id=term_type_id)
    if assess_depart_id and assess_depart_id != 'all':
        query &= Q(assess_depart_id=assess_depart_id)
    if teacher_name:
        query &= Q(teacher__name__icontains=teacher_name)
    if subject_id and subject_id != 'all':
        query &= Q(teacher__subject_id=subject_id)

    # 获取数据
    queryset =PeTeacherMidAssess.objects.filter(query).select_related(
        'semester', 'term_type', 'assess_depart', 'teacher', 'teacher__subject'
    ).order_by('id')

    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "教师期中考核数据"

    # 设置表头
    headers = [
        '序号', '学期', '考核类型', '考核时间', '考核部门', '姓名',
        '教师学科', '课堂节数', '专业课节数折合', '课间操、非工作日学校安排折算', '额外工作折算', '总工作量节数', '工作量成绩', '常规教学薄成绩',
        '总成绩', '名次', '备注', 
    ]

    # 添加表头行
    ws.append(headers)

    # 设置表头样式
    header_font = Font(bold=True, size=10)
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # 添加数据行
    for row_num, obj in enumerate(queryset, start=2):
        data = [
            row_num - 1,
            obj.semester.__str__(),
            obj.term_type.name,
            obj.assess_time,
            obj.assess_depart.name,
            obj.teacher.name,
            obj.teacher.subject.title if obj.teacher.subject else '',
            obj.class_hours,
            obj.major_hours,
            obj.kejiancao_hours,
            obj.extra_work_hours,
            obj.total_workload,
            obj.workload_score,
            obj.teach_book,
            obj.total_score,
            obj.rank,
            obj.remark,           
        ]

        ws.append(data)

        # 设置数据行样式
        for cell in ws[row_num]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 设置文件名
    filename = f"教师期中考核数据_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

    # 准备响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # 保存工作簿到响应
    wb.save(response)

    return response


@superuser_required
def pe_mid_update_rank(request):
    """更新教师期中考核数据的名次并将公示状态改为应经公示"""
    
    if request.method == "POST":
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "请选择Excel文件")
            return redirect('assessments:pe_mid_list')

        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            errors = []
            success_count = 0
            updated_count = 0
            not_found_count = 0

            # 学期映射字典
            semester_map = {}
            for sem in Semester.objects.all():
                key = f"{sem.year}{sem.get_semester_type_display()}"
                semester_map[key] = sem

            # 考核类型映射
            term_type_map = {tt.name: tt for tt in TermType.objects.all()}

            # 教师映射 (姓名->对象)
            teacher_map = {t.name: t for t in UserInfo.objects.all()}

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 跳过空行
                    if not any(row):
                        continue

                    # 获取关键字段
                    semester_str = row[0]
                    term_type_name = row[1]
                    teacher_name = row[2]
                    rank_value = row[3]
                    
                    # 验证必填字段
                    if not all([semester_str, term_type_name, teacher_name, rank_value]):
                        raise ValueError("所有字段都不能为空")
                    
                    # 验证名次格式
                    try:
                        rank_float = float(rank_value)
                        if rank_float <= 0:
                            raise ValueError("名次必须大于0")
                    except (TypeError, ValueError):
                        raise ValueError("名次必须是有效数字")
                    
                    # 获取关联对象
                    semester = semester_map.get(semester_str)
                    if not semester:
                        raise ValueError(f"学期 '{semester_str}' 不存在")
                    
                    term_type = term_type_map.get(term_type_name)
                    if not term_type:
                        raise ValueError(f"考核类型 '{term_type_name}' 不存在")
                    
                    teacher = teacher_map.get(teacher_name)
                    if not teacher:
                        raise ValueError(f"教师 '{teacher_name}' 不存在")
                    
                    # 查找并更新记录
                    try:
                        assess = PeTeacherMidAssess.objects.get(
                            teacher=teacher,
                            semester=semester,
                            term_type=term_type
                        )
                        
                        # 更新名次和公示状态
                        assess.rank = rank_float
                        assess.is_published = True
                        assess.save()
                        
                        success_count += 1
                        updated_count += 1
                        
                    except PeTeacherMidAssess.DoesNotExist:
                        not_found_count += 1
                        errors.append(f"第 {row_num} 行: 找不到匹配的记录 - 教师: {teacher_name}, 学期: {semester_str}, 考核类型: {term_type_name}")

                except Exception as e:
                    errors.append(f"第 {row_num} 行错误: {str(e)}")

            # 结果统计
            result_message = f"成功更新 {success_count} 条记录"
            if not_found_count > 0:
                result_message += f"，未找到 {not_found_count} 条记录"
            
            if errors:
                messages.warning(request, result_message)
                for error in errors[:5]:  # 最多显示5条错误
                    messages.error(request, error)
                if len(errors) > 5:
                    messages.info(request, f"还有 {len(errors)-5} 条错误未显示...")
            else:
                messages.success(request, result_message)

        except Exception as e:
            messages.error(request, f"文件处理错误: {str(e)}")

    return redirect('assessments:pe_mid_list')
