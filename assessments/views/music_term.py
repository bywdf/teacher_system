from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import datetime

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.contrib import messages
from django.db.models import Q
from django.db import transaction

from utils.pagination import Pagination
from utils.bootstrap import BootStrapModelForm

from assessments.models import  MusicTeacherSemesterAssess, AssessDepart, Semester, TermType, MusicTeacherMidAssess, MusicTeacherFinalAssess
from accounts.models import UserInfo, Subject
from utils.user_decorator import superuser_required


class AssessModelForm(BootStrapModelForm):
    class Meta:
        model = MusicTeacherSemesterAssess
        # 字段，所有字段
        fields = '__all__'


def music_term_list(request):
    # 获取所有可选数据
    semesters = Semester.objects.order_by('-id')
    term_types = TermType.objects.all()
    assess_departs = AssessDepart.objects.all()
    subjects = Subject.objects.all()

    # 初始化查询条件
    semester_id = request.GET.get('semester')
    term_type_id = request.GET.get('term_type')
    assess_depart_id = request.GET.get('assess_depart')
    teacher_name = request.GET.get('teacher_name')
    subject_id = request.GET.get('subject')  # 新增学科参数

    # 构建查询条件
    query = Q()
    if semester_id and semester_id != 'all':
        query &= Q(semester_id=semester_id)
    if term_type_id and term_type_id != 'all':
        query &= Q(term_type_id=term_type_id)
    if assess_depart_id and assess_depart_id != 'all':
        query &= Q(assess_depart_id=assess_depart_id)
    if teacher_name:
        query &= Q(teacher__name__icontains=teacher_name)
    if subject_id and subject_id != 'all':
        query &= Q(teacher__subject_id=subject_id)
        
    # 权限控制：普通用户只能查看自己的记录
    if not (request.user.is_superuser or request.user.groups.filter(name='管理员').exists()):
        query &= Q(teacher=request.user)

    # 应用查询条件
    queryset = MusicTeacherSemesterAssess.objects.filter(
        query).order_by('id')

    page_object = Pagination(request, queryset)

    content = {
        'queryset': page_object.page_queryset,
        'page_string': page_object.html(),
        'semesters': semesters,
        'term_types': term_types,
        'assess_departs': assess_departs,
        'subjects': subjects,
        'selected_semester': semester_id if semester_id else 'all',
        'selected_term_type': term_type_id if term_type_id else 'all',
        'selected_assess_depart': assess_depart_id if assess_depart_id else 'all',
        'teacher_name': teacher_name if teacher_name else '',
        'selected_subject': subject_id if subject_id else 'all',

    }
    return render(request, 'music_term_list.html', content)


@superuser_required
def music_term_delete(request):
    """删除"""
    nid = request.GET.get('nid')
    MusicTeacherSemesterAssess.objects.filter(id=nid).delete()
    return redirect('assessments:music_term_list')


@superuser_required
def music_term_edit(request, pk):
    instance = get_object_or_404(MusicTeacherSemesterAssess, pk=pk)
    form = AssessModelForm(request.POST or None, instance=instance)

    if request.method == 'POST':
        if form.is_valid():
            # 保存前自动关联期中期末成绩
            teacher = form.cleaned_data['teacher']
            semester = form.cleaned_data['semester']
            
            # 查找对应的期中成绩
            mid_assess = MusicTeacherMidAssess.objects.filter(
                teacher=teacher, 
                semester=semester
            ).first()
            
            # 查找对应的期末成绩
            final_assess = MusicTeacherFinalAssess.objects.filter(
                teacher=teacher, 
                semester=semester
            ).first()
            
            # 设置关联
            instance.mid_score = mid_assess
            instance.final_score = final_assess
            
            # 计算总成绩
            instance.total_score = (mid_assess.total_score if mid_assess else 0) + \
                                 (final_assess.total_score if final_assess else 0)
            
            instance.save()
            return redirect('assessments:music_term_list')

    context = {
        'form': form,
        'title': '考核记录',
        'instance': instance,
        'show_mid_score': True,
        'show_final_score': True
    }
    return render(request, 'assess_change.html', context)


@superuser_required
def music_term_add(request):
    form = AssessModelForm()
    teachers = UserInfo.objects.all()
    if request.method == 'POST':
        form = AssessModelForm(request.POST)
        if form.is_valid():
            # 保存前自动关联期中期末成绩
            teacher = form.cleaned_data['teacher']
            semester = form.cleaned_data['semester']
            
            # 查找对应的期中成绩
            mid_assess = MusicTeacherMidAssess.objects.filter(
                teacher=teacher, 
                semester=semester
            ).first()
            
            # 查找对应的期末成绩
            final_assess = MusicTeacherFinalAssess.objects.filter(
                teacher=teacher, 
                semester=semester
            ).first()
            
            # 创建实例并设置关联
            instance = form.save(commit=False)
            instance.mid_score = mid_assess
            instance.final_score = final_assess
            instance.total_score = (mid_assess.total_score if mid_assess else 0) + \
                                 (final_assess.total_score if final_assess else 0)
            instance.save()
            
            return redirect('assessments:music_term_list')
    
    return render(request, 'assess_change.html', {'form': form, 'title': '新建考核','show_mid_score': True, 'show_final_score': True})


# 下面是批量导入需要的功能
@superuser_required
@transaction.atomic
def music_term_import(request):
    """批量导入考核成绩并自动关联期中期末成绩"""
    if request.method == "POST":
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "请选择Excel文件")
            return redirect('assessments:music_term_list')

        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            errors = []
            success_count = 0
            created_count = 0
            updated_count = 0

            # 预加载基础数据映射（优化查询性能）
            semester_map = {f"{sem.year}{sem.get_semester_type_display()}": sem 
                           for sem in Semester.objects.all()}
            term_type_map = {tt.name: tt for tt in TermType.objects.all()}
            depart_map = {ad.name: ad for ad in AssessDepart.objects.all()}
            teacher_map = {t.name: t for t in UserInfo.objects.all()}
            
            # 预加载期中期末成绩映射（教师-学期 -> 成绩对象）
            mid_assess_map = {}
            for mid in MusicTeacherMidAssess.objects.all():
                key = (mid.teacher_id, mid.semester_id)
                mid_assess_map[key] = mid
                
            final_assess_map = {}
            for final in MusicTeacherFinalAssess.objects.all():
                key = (final.teacher_id, final.semester_id)
                final_assess_map[key] = final

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 跳过空行
                    if not any(row):
                        continue

                    # 解析必要字段（假设Excel列顺序：学期、考核类型、考核时间、部门、教师、备注）
                    semester_str = row[0].strip() if row[0] else ""
                    term_type_name = row[1].strip() if row[1] else ""
                    assess_time = row[2] if row[2] else datetime.date.today().isoformat()
                    depart_name = row[3].strip() if row[3] else ""
                    teacher_name = row[4].strip() if row[4] else ""
                    remark = row[5] if row[5] else ""

                    # 基础字段校验
                    if not semester_str:
                        raise ValueError("学期不能为空")
                    if not term_type_name:
                        raise ValueError("考核类型不能为空")
                    if not depart_name:
                        raise ValueError("考核部门不能为空")
                    if not teacher_name:
                        raise ValueError("教师姓名不能为空")

                    # 处理学期（支持自动创建不存在的学期）
                    if semester_str not in semester_map:
                        year = semester_str[:-3]  # 假设格式为"2023-2024上学期"，取前部分作为学年
                        semester_type = 'last' if '上' in semester_str else 'next'
                        sem, created = Semester.objects.get_or_create(
                            year=year,
                            semester_type=semester_type
                        )
                        semester_map[semester_str] = sem
                    semester = semester_map[semester_str]

                    # 处理考核类型（自动创建不存在的类型）
                    term_type = term_type_map.get(term_type_name)
                    if not term_type:
                        term_type = TermType.objects.create(name=term_type_name)
                        term_type_map[term_type_name] = term_type

                    # 处理考核部门（自动创建不存在的部门）
                    assess_depart = depart_map.get(depart_name)
                    if not assess_depart:
                        assess_depart = AssessDepart.objects.create(name=depart_name)
                        depart_map[depart_name] = assess_depart

                    # 处理教师（必须已存在）
                    teacher = teacher_map.get(teacher_name)
                    if not teacher:
                        raise ValueError(f"教师 '{teacher_name}' 不存在于系统中")

                    # ------------------- 关键修改：关联期中期末成绩 -------------------
                    # 构建查找键（教师ID+学期ID）
                    key = (teacher.id, semester.id)
                    
                    # 查找期中成绩
                    mid_assess = mid_assess_map.get(key)
                    if not mid_assess:
                        # 尝试通过查询获取（处理映射未加载的情况）
                        mid_assess = MusicTeacherMidAssess.objects.filter(
                            teacher=teacher, 
                            semester=semester
                        ).first()
                        if mid_assess:
                            mid_assess_map[key] = mid_assess  # 更新映射缓存
                    
                    # 查找期末成绩
                    final_assess = final_assess_map.get(key)
                    if not final_assess:
                        final_assess = MusicTeacherFinalAssess.objects.filter(
                            teacher=teacher, 
                            semester=semester
                        ).first()
                        if final_assess:
                            final_assess_map[key] = final_assess  # 更新映射缓存
                    
                    # ------------------- 创建或更新学期总评记录 -------------------
                   # 准备默认值（不包括total_score，让模型自己计算）
                    defaults = {
                        'assess_time': assess_time,
                        'assess_depart': assess_depart,
                        'mid_score': mid_assess,
                        'final_score': final_assess,
                        'remark': remark,
                    }
                    
                    # 获取或创建对象
                    obj, created = MusicTeacherSemesterAssess.objects.get_or_create(
                        teacher=teacher,
                        semester=semester,
                        term_type=term_type,
                        defaults=defaults
                    )
                    
                    # 如果是更新操作，应用更改
                    if not created:
                        for field, value in defaults.items():
                            setattr(obj, field, value)
                    
                    # 确保保存时计算总分
                    obj.save()

                    success_count += 1
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                except Exception as e:
                    errors.append(f"第 {row_num} 行错误: {str(e)}")

            # 处理导入结果反馈
            if errors:
                msg = f"成功导入 {success_count} 条（新增:{created_count}, 更新:{updated_count}），失败 {len(errors)} 条"
                messages.warning(request, msg)
                for error in errors[:5]:  # 最多显示5条错误
                    messages.error(request, error)
                if len(errors) > 5:
                    messages.info(request, f"还有 {len(errors)-5} 条错误未显示...")
            else:
                messages.success(
                    request, f"成功导入 {success_count} 条数据（新增:{created_count}, 更新:{updated_count}）")

        except Exception as e:
            messages.error(request, f"文件处理错误: {str(e)}")

    return redirect('assessments:music_term_list')


def music_term_export(request):
    """导出教师学期考核数据"""
    # 获取筛选参数
    semester_id = request.GET.get('semester')
    term_type_id = request.GET.get('term_type')
    assess_depart_id = request.GET.get('assess_depart')
    teacher_name = request.GET.get('teacher_name')
    subject_id = request.GET.get('subject')

    # 构建查询条件
    query = Q()
    if semester_id and semester_id != 'all':
        query &= Q(semester_id=semester_id)
    if term_type_id and term_type_id != 'all':
        query &= Q(term_type_id=term_type_id)
    if assess_depart_id and assess_depart_id != 'all':
        query &= Q(assess_depart_id=assess_depart_id)
    if teacher_name:
        query &= Q(teacher__name__icontains=teacher_name)
    if subject_id and subject_id != 'all':
        query &= Q(teacher__subject_id=subject_id)

    # 获取数据
    queryset = MusicTeacherSemesterAssess.objects.filter(query).select_related(
        'semester', 'term_type', 'assess_depart', 'teacher', 'teacher__subject'
    ).order_by('id')

    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "教师学期考核数据"

    # 设置表头
    headers = [
        '序号', '学期', '考核类型', '考核时间', '考核部门', '姓名','教师学科', 
        '期中成绩', '期末成绩', '学期成绩', '名次', '备注'    
    ]

    # 添加表头行
    ws.append(headers)

    # 设置表头样式
    header_font = Font(bold=True, size=10)
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # 添加数据行
    for row_num, obj in enumerate(queryset, start=2):
        data = [
            row_num - 1,
            obj.semester.__str__(),
            obj.term_type.name,
            obj.assess_time,
            obj.assess_depart.name,
            obj.teacher.name,
            obj.teacher.subject.title if obj.teacher.subject else '',
            obj.mid_score.total_score if obj.mid_score else '',
            obj.final_score.total_score if obj.final_score else '',
            obj.total_score,
            obj.rank,
            obj.remark,       
        ]

        ws.append(data)

        # 设置数据行样式
        for cell in ws[row_num]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 设置文件名
    filename = f"教师学期考核数据_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

    # 准备响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # 保存工作簿到响应
    wb.save(response)

    return response


@superuser_required
def music_term_update_rank(request):
    """更新教师学期考核数据的名次并将公示状态改为应经公示"""
    
    if request.method == "POST":
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "请选择Excel文件")
            return redirect('assessments:music_term_list')

        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            errors = []
            success_count = 0
            updated_count = 0
            not_found_count = 0

            # 学期映射字典
            semester_map = {}
            for sem in Semester.objects.all():
                key = f"{sem.year}{sem.get_semester_type_display()}"
                semester_map[key] = sem

            # 考核类型映射
            term_type_map = {tt.name: tt for tt in TermType.objects.all()}

            # 教师映射 (姓名->对象)
            teacher_map = {t.name: t for t in UserInfo.objects.all()}

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 跳过空行
                    if not any(row):
                        continue

                    # 获取关键字段
                    semester_str = row[0]
                    term_type_name = row[1]
                    teacher_name = row[2]
                    rank_value = row[3]
                    
                    # 验证必填字段
                    if not all([semester_str, term_type_name, teacher_name, rank_value]):
                        raise ValueError("所有字段都不能为空")
                    
                    # 验证名次格式
                    try:
                        rank_float = float(rank_value)
                        if rank_float <= 0:
                            raise ValueError("名次必须大于0")
                    except (TypeError, ValueError):
                        raise ValueError("名次必须是有效数字")
                    
                    # 获取关联对象
                    semester = semester_map.get(semester_str)
                    if not semester:
                        raise ValueError(f"学期 '{semester_str}' 不存在")
                    
                    term_type = term_type_map.get(term_type_name)
                    if not term_type:
                        raise ValueError(f"考核类型 '{term_type_name}' 不存在")
                    
                    teacher = teacher_map.get(teacher_name)
                    if not teacher:
                        raise ValueError(f"教师 '{teacher_name}' 不存在")
                    
                    # 查找并更新记录
                    try:
                        assess = MusicTeacherSemesterAssess.objects.get(
                            teacher=teacher,
                            semester=semester,
                            term_type=term_type
                        )
                        
                        # 更新名次和公示状态
                        assess.rank = rank_float
                        assess.is_published = True
                        assess.save()
                        
                        success_count += 1
                        updated_count += 1
                        
                    except MusicTeacherSemesterAssess.DoesNotExist:
                        not_found_count += 1
                        errors.append(f"第 {row_num} 行: 找不到匹配的记录 - 教师: {teacher_name}, 学期: {semester_str}, 考核类型: {term_type_name}")

                except Exception as e:
                    errors.append(f"第 {row_num} 行错误: {str(e)}")

            # 结果统计
            result_message = f"成功更新 {success_count} 条记录"
            if not_found_count > 0:
                result_message += f"，未找到 {not_found_count} 条记录"
            
            if errors:
                messages.warning(request, result_message)
                for error in errors[:5]:  # 最多显示5条错误
                    messages.error(request, error)
                if len(errors) > 5:
                    messages.info(request, f"还有 {len(errors)-5} 条错误未显示...")
            else:
                messages.success(request, result_message)

        except Exception as e:
            messages.error(request, f"文件处理错误: {str(e)}")

    return redirect('assessments:music_term_list')
