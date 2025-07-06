from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from openpyxl import load_workbook

from assessments.models import AssessDepart
from utils.bootstrap import BootStrapModelForm
from utils.pagination import Pagination
from utils.user_decorator import superuser_required, admin_or_superuser_required

# Create your views here.


class AssessDepartModelForm(BootStrapModelForm):
    class Meta:
        model = AssessDepart
        fields = '__all__'


@admin_or_superuser_required
def assessdepart_list(request):
    '''考核部门列表'''
    form = AssessDepartModelForm()
    query = AssessDepart.objects.all()
    page_object = Pagination(request, query)
    context = {
        'form': form,
        'page_string': page_object.html(),
        'queryset': page_object.page_queryset,
    }
    return render(request, 'assessdepart_list.html', context)



@csrf_exempt
@superuser_required
def  assessdepart_add(request):
    '''考核部门添加(Ajax请求)'''
    form = AssessDepartModelForm(request.POST)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({'status': False, 'error': form.errors})


@superuser_required
def assessdepart_delete(request):
    '''考核部门删除'''
    uid = request.GET.get('uid')
    exists = AssessDepart.objects.filter(id=uid).exists()
    if not exists:
        return JsonResponse({'status': False, 'error': '删除失败，数据不存在'})
    AssessDepart.objects.filter(id=uid).delete()
    return JsonResponse({'status': True})


def assessdepart_detail(request):
    '''根据ID获取考核部门详情'''
    uid = request.GET.get('uid')
    # row_dict = models.assessdepart.objects.filter(id=uid).values('title').first()
    row_dict = AssessDepart.objects.filter(id=uid).values('name').first()
    if not row_dict:
        return JsonResponse({"status":False, 'error':"数据不存在"})
    result = {
        'status': True,
        'data': row_dict,
    }
    return JsonResponse(result)
    
@csrf_exempt
@superuser_required
def assessdepart_edit(request):
    '''编辑考核部门'''
    uid = request.GET.get('uid')
    row_object = AssessDepart.objects.filter(id=uid).first()
    if not row_object:
        return JsonResponse({"status":False, 'error':"数据不存在,请刷新重试"})

    form = AssessDepartModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({'status': False, 'error': form.errors})


@superuser_required
def assessdepart_import(request):
    '''Excel表格批量导入考核部门'''
    # django.core.files.uploadedfile.InMemoryUploadedFile

    # 1.获取用户上传的对象文件
    file_object = request.FILES.get("exc")
    # print(type(file_object))

    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value

        exists = AssessDepart.objects.filter(name=text).exists()
        if not exists:
            AssessDepart.objects.create(name=text)

    # with open(file_object.title, mode='wb') as f:
    #     for chunk in file_object:
    #         f.write(chunk)
    return redirect("assessments:assessdepart_list")