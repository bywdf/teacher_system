from openpyxl import load_workbook
from assessments.forms import BatchImportForm
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from assessments.models import TeacherMidAssess, AssessDepart, Semester, TermType

from utils.pagination import Pagination
from utils.bootstrap import BootStrapModelForm

from accounts.models import UserInfo
from django.db.models import Q, F
from django.http import JsonResponse
from django.db import transaction
import datetime


def teacher_autocomplete(request):
    query = request.GET.get('q', '')
    # 根据姓名或用户名模糊查询教师
    teachers = UserInfo.objects.filter(
        Q(name__icontains=query) | Q(username__icontains=query)
    ).values('id', 'name', 'username')[:10]  # 最多返回10条结果
    return JsonResponse(list(teachers), safe=False)


class MidAssessModelForm(BootStrapModelForm):
    class Meta:
        model = TeacherMidAssess
        # 排除字段
        fields = '__all__'


def cultura_mid_list(request):
    queryset = TeacherMidAssess.objects.all().order_by('id')
    page_object = Pagination(request, queryset)
    content = {
        'queryset': page_object.page_queryset,
        'page_string': page_object.html(),
    }
    return render(request, 'cultura_mid_list.html', content)


def cultura_mid_delete(request):
    """删除"""
    nid = request.GET.get('nid')
    TeacherMidAssess.objects.filter(id=nid).delete()
    return redirect('assessments:cultura_mid_list')


def cultura_mid_edit(request, pk):
    # 获取要编辑的对象，若不存在则返回404
    instance = get_object_or_404(TeacherMidAssess, pk=pk)
    # 创建表单实例，绑定现有数据
    form = MidAssessModelForm(request.POST or None, instance=instance)

    if request.method == 'POST':
        if form.is_valid():
            # 保存更新（可在此处添加额外逻辑，如权限检查、计算字段等）
            form.save()
            return redirect('assessments:cultura_mid_list')  # 重定向到列表页
        # 若表单验证失败，保留错误信息并重新渲染页面

    # 渲染编辑页面，传递表单和对象
    context = {
        'form': form,
        'title': '编辑考核记录',
        'instance': instance
    }
    return render(request, 'assess_change.html', context)


def cultura_mid_add(request):
    """添加"""
    form = MidAssessModelForm()
    # 获取所有教师数据
    teachers = UserInfo.objects.all()
    if request.method == 'POST':
        form = MidAssessModelForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('assessments:cultura_mid_list')
    return render(request, 'assess_change.html', {'form': form})


# 下面是批量导入需要的功能


@transaction.atomic
def cultura_mid_import(request):
    if request.method == "POST":
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "请选择Excel文件")
            return redirect('assessments:cultura_mid_list')
        
        try:
            wb = load_workbook(excel_file)
            ws = wb.active
            errors = []
            success_count = 0
            created_count = 0
            updated_count = 0
            
            # 学期映射字典
            semester_map = {}
            for sem in Semester.objects.all():
                key = f"{sem.year}{sem.get_semester_type_display()}"
                semester_map[key] = sem
            
            # 考核类型映射
            term_type_map = {tt.name: tt for tt in TermType.objects.all()}
            
            # 部门映射
            depart_map = {ad.name: ad for ad in AssessDepart.objects.all()}
            
            # 教师映射 (姓名->对象)
            teacher_map = {t.name: t for t in UserInfo.objects.all()}
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 跳过空行
                    if not any(row):
                        continue
                    
                    # 解析学期 (格式: "2023-2024上学期")
                    semester_str = row[0]
                    if not semester_str:
                        raise ValueError("学期不能为空")
                    
                    if semester_str not in semester_map:
                        # 尝试创建新学期
                        year = semester_str[:-3]  # 去掉后3个字符(学期名)
                        semester_type = 'last' if '上' in semester_str else 'next'
                        sem, created = Semester.objects.get_or_create(
                            year=year,
                            semester_type=semester_type
                        )
                        semester_map[semester_str] = sem
                    
                    # 获取其他关联对象
                    term_type_name = row[1]
                    if not term_type_name:
                        raise ValueError("考核类型不能为空")
                    
                    term_type = term_type_map.get(term_type_name)
                    if not term_type:
                        term_type = TermType.objects.create(name=term_type_name)
                        term_type_map[term_type_name] = term_type
                    
                    depart_name = row[3]
                    if not depart_name:
                        raise ValueError("考核部门不能为空")
                    
                    assess_depart = depart_map.get(depart_name)
                    if not assess_depart:
                        assess_depart = AssessDepart.objects.create(name=depart_name)
                        depart_map[depart_name] = assess_depart
                    
                    teacher_name = row[4]
                    if not teacher_name:
                        raise ValueError("教师姓名不能为空")
                    
                    teacher = teacher_map.get(teacher_name)
                    if not teacher:
                        raise ValueError(f"教师 '{teacher_name}' 不存在")
                    
                    # 转换字段值
                    def safe_float(value, default=0.0):
                        try:
                            return float(value) if value is not None else default
                        except (TypeError, ValueError):
                            return default
                    
                    # 创建或更新记录
                    obj, created = TeacherMidAssess.objects.update_or_create(
                        teacher=teacher,
                        semester=semester_map[semester_str],
                        term_type=term_type,
                        defaults={
                            'assess_time': row[2] or datetime.date.today().isoformat(),
                            'assess_depart': assess_depart,
                            'class_hours': safe_float(row[5]),
                            'duty_hours': safe_float(row[6]),
                            'extra_work_hours': safe_float(row[7]),
                            'personal_score': safe_float(row[8]),
                            'class_score': safe_float(row[9]),
                            'group_score': safe_float(row[10]),
                            'remark': row[11] or "",
                            'week': int(row[12]) if len(row) > 12 and row[12] is not None else 10
                        }
                    )
                    
                    success_count += 1
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
                    
                except Exception as e:
                    errors.append(f"第 {row_num} 行错误: {str(e)}")
            
            if errors:
                messages.warning(request, f"成功导入 {success_count} 条（新增:{created_count}, 更新:{updated_count}），失败 {len(errors)} 条")
                for error in errors[:5]:  # 最多显示5条错误
                    messages.error(request, error)
                if len(errors) > 5:
                    messages.info(request, f"还有 {len(errors)-5} 条错误未显示...")
            else:
                messages.success(request, f"成功导入 {success_count} 条数据（新增:{created_count}, 更新:{updated_count}）")
                
        except Exception as e:
            messages.error(request, f"文件处理错误: {str(e)}")
    
    return redirect('assessments:cultura_mid_list') 