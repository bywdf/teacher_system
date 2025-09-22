from django.shortcuts import render

from certificate.models import SchoolCertificate
from certificate import models
from django.shortcuts import redirect
from utils.bootstrap import BootStrapModelForm
from utils.pagination import Pagination
from utils.user_decorator import admin_or_superuser_required, superuser_required


class SchoolCertificateModelForm(BootStrapModelForm):
    class Meta:
        model = SchoolCertificate
        fields = '__all__'
    

@admin_or_superuser_required
def school_certificate_list(request):
    """学校证书"""
    queryset = SchoolCertificate.objects.all()
    page_object = Pagination(request, queryset)
    context = {
        'queryset': page_object.page_queryset,     # 分页数据
        'page_string': page_object.html()         # 生成页码
    }
    
    return render(request, 'school_certificate.html', context)

@superuser_required
def school_certificate_add(request):
    '''添加校级证书'''
    if request.method == 'GET':
        return render(request, 'school_certificate_add.html')

    name = request.POST.get('name')
    models.SchoolCertificate.objects.create(name=name)
    return redirect('/certificate/school_certificate/list/')


@superuser_required
def school_certificate_delete(request):
    '''删除校级证书'''
    nid = request.GET.get('nid')
    models.SchoolCertificate.objects.filter(id=nid).delete()
    return redirect('/certificate/school_certificate/list/')


@superuser_required
def school_certificate_edit(request, nid):
    '''修改校级证书'''
    # 通过nid获取数据，获取到的是一个列表对象，获取第一个
    if request.method == 'GET':
        row_object = models.SchoolCertificate.objects.filter(id=nid).first()
        return render(request, 'school_certificate_edit.html', {'row_object': row_object})
    name = request.POST.get('name')
    models.SchoolCertificate.objects.filter(id=nid).update(name=name)
    return redirect('/certificate/school_certificate/list')


@superuser_required
def school_certificate_multi(request):
    '''批量上传 (excel)'''
    from openpyxl import load_workbook
    # django.core.files.uploadedfile.InMemoryUploadedFile

    # 1.获取用户上传的对象文件
    file_object = request.FILES.get("exc")
    # print(type(file_object))

    # 2.对象传递给openpyxl，由openpyxl读取文件的内容
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    # 3.循环获取每一行数据
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value

        exists = models.SchoolCertificate.objects.filter(title=text).exists()
        if not exists:
            models.SchoolCertificate.objects.create(title=text)

    # with open(file_object.title, mode='wb') as f:
    #     for chunk in file_object:
    #         f.write(chunk)

    return redirect("/certificate/school_certificate/list/")
